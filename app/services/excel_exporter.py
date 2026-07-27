from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from sqlmodel import Session, select

from app.models.analysis_result import AnalysisResult
from app.models.document import Document
from app.models.drawing import Drawing


class ExcelExporter:

    def __init__(
        self,
        session: Session,
    ):
        self.session = session

    def export(
        self,
        document: Document,
    ) -> BytesIO:

        workbook = Workbook()

        sheet = workbook.active
        sheet.title = "Analysis"

        headers = [
            "Страница",
            "Наименование",
            "Количество",
            "Ширина",
            "Высота",
            "Количество дверей",
        ]

        fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3",
        )

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        bold = Font(
            bold=True,
        )

        for column, header in enumerate(headers, start=1):

            cell = sheet.cell(
                row=1,
                column=column,
                value=header,
            )

            cell.fill = fill
            cell.font = bold
            cell.border = border
            cell.alignment = alignment

        statement = (
            select(
                Drawing,
                AnalysisResult,
            )
            .join(
                AnalysisResult,
            )
            .where(
                Drawing.document_id == document.id,
            )
            .order_by(
                Drawing.page_number,
                Drawing.drawing_number,
            )
        )

        row = 2

        total_square = 0
        total_doors = 0

        for drawing, result in self.session.exec(statement):

            values = [
                drawing.page_number,
                result.name,
                result.quantity,
                result.width,
                result.height,
                result.doors_count,
            ]

            for column, value in enumerate(values, start=1):

                cell = sheet.cell(
                    row=row,
                    column=column,
                    value=value,
                )

                cell.border = border
                cell.alignment = alignment

            total_square += (
                result.quantity *
                result.width *
                result.height
            )

            total_doors += result.doors_count*result.quantity

            row += 1

        # Пустая строка
        row += 1

        # Итоговая площадь
        sheet.cell(
            row=row,
            column=1,
            value="Общая площадь, м²"
        ).font = bold

        sheet.cell(
            row=row,
            column=2,
            value=round(total_square * 1e-6, 2),
        ).font = bold


        # Следующая строка
        row += 1

        # Итоговое количество дверей
        sheet.cell(
            row=row,
            column=1,
            value="Кол-во дверей, шт."
        ).font = bold

        sheet.cell(
            row=row,
            column=2,
            value=total_doors,
        ).font = bold

        # Оформление итоговых строк
        for r in (row - 1, row):
            for c in range(1, 3):
                cell = sheet.cell(row=r, column=c)
                cell.border = border
                cell.alignment = alignment

        widths = {
            "A": 20,
            "B": 15,
            "D": 40,
            "C": 15,
            "D": 15,
            "E": 15,
            "F": 18,
        }

        for column, width in widths.items():

            sheet.column_dimensions[column].width = width

        output = BytesIO()

        workbook.save(output)
        workbook.close()

        output.seek(0)

        return output